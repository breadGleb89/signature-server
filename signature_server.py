import os
import base64
import uuid
import threading
import requests
import time
import json
from io import BytesIO
from datetime import datetime
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from PIL import Image

app = Flask(__name__)
CORS(app)

# Папки для хранения
SIGNATURES_FOLDER = "signatures"
FORMS_FOLDER = "forms"
os.makedirs(SIGNATURES_FOLDER, exist_ok=True)
os.makedirs(FORMS_FOLDER, exist_ok=True)

# Конфигурация Telegram бота
BOT_TOKEN = "8629592192:AAFxqvsVSGClASESKVNqLSAD4qEI3rQ98n4"
ADMIN_ID = 454939631

# Keep-alive функция
def keep_alive():
    url = "https://signature-server-87mz.onrender.com/health"
    while True:
        time.sleep(600)
        try:
            requests.get(url, timeout=5)
            print("✅ Keep-alive ping отправлен")
        except:
            print("❌ Ошибка пинга")

threading.Thread(target=keep_alive, daemon=True).start()

# =====================
# ФУНКЦИИ ДЛЯ EXCEL
# =====================
def fill_excel_data(wb, data):
    from openpyxl import load_workbook
    ws_main = wb.active
    
    ws_main["B3"] = data.get("name", "")
    ws_main["E2"] = data.get("order_date", "")
    ws_main["C19"] = data.get("sender_address", "")
    ws_main["C23"] = data.get("polname", "")
    ws_main["C21"] = f"г.{data.get('city', '')}, ул. {data.get('street', '')}, {data.get('house', '')}"
    ws_main["D5"] = data.get("sender_phone", "")
    ws_main["D25"] = data.get("phone", "")
    ws_main["G7"] = data.get("passport_number", "")
    ws_main["B8"] = data.get("issued_by", "")
    ws_main["G9"] = data.get("passport_date", "")
    ws_main["C10"] = data.get("registration", "")
    ws_main["D14"] = data.get("type", "")
    ws_main["B16"] = data.get("weight", "")
    ws_main["C18"] = data.get("places", "")
    
    # Паспортная серия
    passport_series = data.get("passport_series", "")
    if passport_series:
        for merged in ws_main.merged_cells.ranges:
            if "E7" in str(merged):
                ws_main[merged.start_cell.coordinate] = passport_series
                break
    
    if "Опись" in wb.sheetnames:
        ws = wb["Опись"]
        ws["A9"] = f" г. {data.get('city', '')} , {data.get('polname', '')} , {data.get('phone', '')}"
        ws["E7"] = f"  {data.get('sender_address', '')} , {data.get('sender_contry', '')} в "
        ws["E49"] = data.get("places", "")
    
    return wb

def write_items_to_excel(wb, items, start_row=14):
    from openpyxl.styles import Border, Side, Alignment, Font
    
    if "Опись" not in wb.sheetnames:
        wb.create_sheet("Опись")
    
    ws = wb["Опись"]
    
    total_row = 43
    for row in range(start_row, 50):
        try:
            if ws.cell(row=row, column=2).value == "ИТОГО:":
                total_row = row
                break
        except:
            continue
    
    for row in range(start_row, total_row):
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            try:
                ws[f"{col}{row}"].value = None
            except:
                pass
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for i, item in enumerate(items):
        row = start_row + i
        if row >= total_row:
            break
        
        try:
            qty = float(str(item.get('qty', '0')).replace(',', '.'))
            price = float(str(item.get('price', '0')).replace(',', '.'))
        except:
            qty = float(item.get('qty', 0))
            price = float(item.get('price', 0))
        
        total = qty * price
        price_str = str(price).replace('.', ',')
        
        ws[f"A{row}"] = i + 1
        ws[f"B{row}"] = item.get('name', '')
        ws[f"C{row}"] = item.get('unit', '')
        ws[f"D{row}"] = qty
        ws[f"E{row}"] = price_str
        ws[f"F{row}"] = total
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            cell = ws[f"{col}{row}"]
            cell.border = thin_border
            if col == 'A':
                cell.alignment = Alignment(horizontal='left')
    
    if ws.cell(row=total_row, column=2).value != "ИТОГО:":
        ws.cell(row=total_row, column=2, value="ИТОГО:")
        ws.cell(row=total_row, column=2).font = Font(bold=True)
    
    return wb

# =====================
# ФУНКЦИИ ДЛЯ ОТПРАВКИ В TELEGRAM
# =====================
def send_file_to_telegram(chat_id, filepath, filename):
    """Отправляет файл в Telegram"""
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument"
    
    with open(filepath, 'rb') as f:
        files = {'document': (filename, f)}
        data = {'chat_id': chat_id}
        response = requests.post(url, files=files, data=data)
    
    return response.json()

def send_message_to_telegram(chat_id, text):
    """Отправляет сообщение в Telegram"""
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
    data = {'chat_id': chat_id, 'text': text, 'parse_mode': 'Markdown'}
    response = requests.post(url, json=data)
    return response.json()

# =====================
# ФУНКЦИИ ДЛЯ ПОДПИСЕЙ
# =====================
@app.route('/save_signature', methods=['POST', 'OPTIONS'])
def save_signature():
    if request.method == 'OPTIONS':
        return jsonify({}), 200
    
    try:
        data = request.json
        image_data = data.get('signature')
        user_id = data.get('user_id', 'unknown')
        
        if not image_data:
            return jsonify({'error': 'No image'}), 400
        
        if ',' in image_data:
            base64_data = image_data.split(',')[1]
        else:
            base64_data = image_data
        
        image_bytes = base64.b64decode(base64_data)
        
        img = Image.open(BytesIO(image_bytes))
        img.thumbnail((300, 150), Image.Resampling.LANCZOS)
        
        if img.mode in ('RGBA', 'LA', 'P'):
            background = Image.new('RGB', img.size, (255, 255, 255))
            if img.mode == 'P':
                img = img.convert('RGBA')
            if img.mode == 'RGBA':
                background.paste(img, mask=img.split()[-1])
            else:
                background.paste(img)
            img = background
        elif img.mode != 'RGB':
            img = img.convert('RGB')
        
        unique_id = str(uuid.uuid4())[:8]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"signature_{user_id}_{timestamp}_{unique_id}.jpg"
        filepath = os.path.join(SIGNATURES_FOLDER, filename)
        img.save(filepath, 'JPEG', quality=70, optimize=True)
        
        base_url = os.environ.get('BASE_URL', 'https://signature-server-87mz.onrender.com')
        image_url = f"{base_url}/get_signature/{filename}"
        
        return jsonify({'status': 'ok', 'url': image_url})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/get_signature/<filename>', methods=['GET'])
def get_signature(filename):
    filepath = os.path.join(SIGNATURES_FOLDER, filename)
    if os.path.exists(filepath):
        return send_file(filepath, mimetype='image/jpeg')
    return jsonify({'error': 'File not found'}), 404

# =====================
# ГЛАВНАЯ ФУНКЦИЯ - СОЗДАНИЕ ФАЙЛОВ И ОТПРАВКА В TELEGRAM
# =====================
@app.route('/submit_form', methods=['POST', 'OPTIONS'])
def submit_form():
    """Получает данные формы, создает Excel и отправляет в Telegram"""
    if request.method == 'OPTIONS':
        return jsonify({}), 200
    
    try:
        data = request.json
        form_data = data.get('form_data')
        user_telegram_id = data.get('user_id')
        
        if not form_data:
            return jsonify({'error': 'No form data'}), 400
        
        print(f"📝 Получена форма от пользователя {user_telegram_id}")
        print(f"👤 Заказчик: {form_data.get('name')}")
        
        # 1. Скачиваем подпись
        signature_url = form_data.get('signature_url')
        signature_bytes = None
        
        if signature_url:
            response = requests.get(signature_url, timeout=30)
            if response.status_code == 200:
                signature_bytes = response.content
                print("✅ Подпись загружена")
        
        # 2. Создаем Excel файл
        from openpyxl import load_workbook
        
        wb = load_workbook("template.xlsx")
        wb = fill_excel_data(wb, form_data)
        wb = write_items_to_excel(wb, form_data.get("items", []), 14)
        
        # Добавляем подпись
        if signature_bytes:
            try:
                from openpyxl.drawing.image import Image as XLImage
                img = XLImage(BytesIO(signature_bytes))
                img.width = 70
                img.height = 30
                wb.active.add_image(img, "G49")
                
                if "Опись" in wb.sheetnames:
                    ws_opis = wb["Опись"]
                    items_count = len(form_data.get("items", []))
                    target_row = 50
                    if items_count > 36:
                        target_row = 50 + (items_count - 36)
                    img2 = XLImage(BytesIO(signature_bytes))
                    img2.width = 70
                    img2.height = 30
                    ws_opis.add_image(img2, f"F{target_row}")
            except Exception as e:
                print(f"Ошибка добавления подписи: {e}")
        
        # Сохраняем файл
        safe_name = re.sub(r'[<>:"/\\|?*]', '_', form_data.get('name', 'unknown'))
        filename = f"{safe_name}_{datetime.now().strftime('%d-%m-%Y')}_{form_data.get('phone', '000')}.xlsx"
        filepath = os.path.join(FORMS_FOLDER, filename)
        wb.save(filepath)
        wb.close()
        
        print(f"✅ Excel файл создан: {filename}")
        
        # 3. Отправляем файл пользователю в Telegram
        if user_telegram_id:
            send_message_to_telegram(user_telegram_id, "📊 *Ваш заказ-поручение готов!*\n\nФайл формируется...")
            
            result = send_file_to_telegram(user_telegram_id, filepath, filename)
            
            if result.get('ok'):
                print(f"✅ Файл отправлен пользователю {user_telegram_id}")
                send_message_to_telegram(
                    user_telegram_id,
                    "✅ *Готово!*\n\n"
                    "📄 Документ во вложении.\n\n"
                    "🙏 Спасибо за использование бота!"
                )
            else:
                print(f"❌ Ошибка отправки: {result}")
        
        # 4. Отправляем копию админу
        send_message_to_telegram(ADMIN_ID, f"📦 *Новый заказ!*\n\n👤 {form_data.get('name')}\n📍 {form_data.get('city')}\n📞 {form_data.get('phone')}")
        send_file_to_telegram(ADMIN_ID, filepath, filename)
        
        # 5. Удаляем временный файл
        os.remove(filepath)
        
        # 6. Возвращаем ответ WebApp
        base_url = os.environ.get('BASE_URL', 'https://signature-server-87mz.onrender.com')
        file_url = f"{base_url}/get_excel/{filename}"
        
        return jsonify({
            'status': 'ok',
            'message': 'Файлы отправлены в Telegram',
            'file_url': file_url
        })
        
    except Exception as e:
        print(f"❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/get_excel/<filename>', methods=['GET'])
def get_excel(filename):
    filepath = os.path.join(FORMS_FOLDER, filename)
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True, download_name=filename)
    return jsonify({'error': 'File not found'}), 404

@app.route('/health', methods=['GET'])
def health():
    return jsonify({
        'status': 'ok',
        'signatures': len(os.listdir(SIGNATURES_FOLDER)),
        'forms': len(os.listdir(FORMS_FOLDER))
    })

if __name__ == '__main__':
    import re
    port = int(os.environ.get('PORT', 10000))
    print(f"🚀 Сервер запущен на порту {port}")
    app.run(host='0.0.0.0', port=port)
